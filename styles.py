# styles.py
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Fonts
header_white_font = Font(name='Calibri', bold=True, color="FFFFFF")
header_black_font = Font(name='Calibri', bold=True, color="000000")
text_font = Font(name='Calibri', color="000000")
white_text_font = Font(name='Calibri', color="FFFFFF")

# Fills
header_black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
header_grey_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
header_dniszkolne_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
header_wakacje_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

# Alignments
alignment_center = Alignment(horizontal="center", vertical="center")
alignment_right = Alignment(horizontal="right", vertical="center")
alignment_left_corner = Alignment(horizontal="left", vertical="top")
alignment_normal = Alignment(horizontal="left", vertical="center")
alignment_hour = Alignment(wrap_text=True, horizontal="center", vertical="center")

# Borders
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

thick_border = Border(
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'),
    bottom=Side(style='thick')
)
