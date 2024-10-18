import openpyxl
from datetime import datetime
from openpyxl.utils import get_column_letter
from styles import*

# Ścieżka do pliku Excel z ktorego pobierane sa dane do rozkladu
sciezka_do_pliku = r"C:\Users\matwa\Desktop\A.xlsx"

# Otwieranie pliku Excel
wb = openpyxl.load_workbook(sciezka_do_pliku)

# Wybór arkusza (możesz podać nazwę arkusza lub użyć domyślnego)
sheet = wb.active

# Tablica do przechowywania numerów kolumn z kursami, które kursują w danym typie dnia
dni_robocze_szkolne = []
wakacje = []
soboty = []
niedziele = []

data_waznosci="01.09.2024"

# Przejście po pierwszym wierszu (nagłówki), (odczytanie w jakie dni kursuje dany kurs)
for idx, cell in enumerate(sheet[1], start=1):  # Przechodzimy po wierszu 1
    if cell.value in {"D", "E", "S"}:  # Sprawdzamy, czy wartość komórki to "D"
        dni_robocze_szkolne.append(idx)  # Zapisujemy numer kolumny (idx)
    if cell.value in {"H", "D"}:  # Sprawdzamy, czy wartość komórki to "D"
        wakacje.append(idx)  # Zapisujemy numer kolumny (idx)
    if cell.value in {"6", "E"}:  # Sprawdzamy, czy wartość komórki to "D"
        soboty.append(idx)  # Zapisujemy numer kolumny (idx)



# Pobierz przystanki z pierwszej kolumny (od wiersza 2 do końca, zakładając że wiersz 1 to nagłówki), w praktyce lista wszystkich przystanków z tabeli
przystanki = [cell[0] for cell in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True)]

# Zbiór unikalnych wariantów
warianty = {}

# Funkcja przypisująca wariant
def przypisz_wariant(przystanki_wariantowe, nazwa_wariantu_max):
    for istniejący_wariant, przystanki_w_wariancie in warianty.items():
        if set(przystanki_wariantowe) == set(przystanki_w_wariancie):
            return istniejący_wariant, nazwa_wariantu_max  # Zwracamy oba elementy
    
    # Jeśli to nowy zestaw przystanków, utwórz nowy wariant
    nowy_wariant = nazwa_wariantu_max
    
    # Dodaj nowy wariant do słownika
    warianty[nowy_wariant] = przystanki_wariantowe

    # Zwiększ literę wariantu (np. A -> B -> C)
    nazwa_wariantu_max = chr(ord(nazwa_wariantu_max) + 1)  # Zwiększ literę (A, B, C...)
    
    return nowy_wariant, nazwa_wariantu_max  # Zwracamy nowy wariant i zaktualizowaną nazwę


# Funkcja obliczająca różnicę czasu między dwoma przystankami
def oblicz_roznice_czasu(start_time, end_time):
    fmt = "%H:%M"
    try:
        time1 = datetime.strptime(start_time, fmt)
        time2 = datetime.strptime(end_time, fmt)
        delta = time2 - time1
        return delta.total_seconds() / 60  # Zwraca różnicę w minutach
    except ValueError:
        return None  # Jeśli wartości nie są prawidłowe, zwróć None

# Funkcja sprawdzająca wariant i obliczająca czas między przystankami
# Funkcja sprawdzająca wariant i obliczająca czas między przystankami
def sprawdz_wariant_od_wiersza(sheet, przystanki, wiersz_startowy):
    warianty_kursow = []  # Lista na wyniki dla każdego kursu
    nazwa_wariantu_max = 'A'  # Resetowanie nazwy wariantu do 'A'

    # Przejście po każdej kolumnie (od kolumny 2 do końca, czyli sprawdzamy przystanki)
    for col_idx, col in enumerate(sheet.iter_cols(min_col=2, max_col=sheet.max_column, min_row=wiersz_startowy, values_only=True), start=2):
        # Sprawdź, czy komórka w wierszu startowym jest pusta
        cell_value = sheet.cell(row=wiersz_startowy, column=col_idx).value
        if cell_value is None or str(cell_value).strip() == "":
            continue  # Pomiń tę kolumnę, jeśli komórka w wierszu startowym jest pusta

        kurs_zatrzymywane_przystanki = []  # Zbieramy przystanki, na których autobus się zatrzymuje
        czas_przejazdu_miedzy_przystankami = []  # Lista na różnice czasu między przystankami

        poprzedni_czas = None  # Zmienna do przechowywania czasu z poprzedniego przystanku
        czas=0
        # Przejście po wierszach w danej kolumnie, zaczynając od wiersza startowego
        for row_idx, value in enumerate(col, start=wiersz_startowy):
            if value is not None and str(value).strip() != "":  # Zbieramy przystanki, gdzie autobus się zatrzymuje
                przystanek = przystanki[row_idx - 2]  # Zapisujemy przystanek, gdzie autobus się zatrzymuje
                kurs_zatrzymywane_przystanki.append(przystanek)

                # Obliczamy różnicę czasu, jeśli poprzedni przystanek był obsługiwany
                if poprzedni_czas is not None:
                    roznica_czasu = oblicz_roznice_czasu(poprzedni_czas, str(value))
                    czas+=roznica_czasu
                    czas_przejazdu_miedzy_przystankami.append({
                        "od_przystanku": przystanki[row_idx - 3],
                        "do_przystanku": przystanek,
                        "czas_przejazdu": czas
                    })

                poprzedni_czas = str(value)  # Aktualizujemy czas na bieżący przystanek

        # Przypisujemy wariant na podstawie zatrzymywanych przystanków
        wariant, nazwa_wariantu_max = przypisz_wariant(kurs_zatrzymywane_przystanki, nazwa_wariantu_max)
        # Zapisz wariant oraz czas przejazdu dla danej kolumny (kursu)
        warianty_kursow.append({
            "kolumna": col_idx,  # Numer kolumny kursu
            "wariant": wariant,  # Wariant przypisany do kursu
            "zatrzymywane_przystanki": kurs_zatrzymywane_przystanki,  # Lista przystanków, gdzie autobus się zatrzymuje
            "czas_przejazdu_miedzy_przystankami": czas_przejazdu_miedzy_przystankami  # Lista z czasami przejazdów między przystankami
        })

    return warianty_kursow




def usun_przystanki_bez_kursow(przystanki, warianty_kursow, numer_wiersza):
    # Tworzymy słownik, który będzie przechowywał informację o tym, czy przystanek jest obsługiwany przez jakiś kurs
    index = numer_wiersza - 2
    przystanki1 = przystanki[index:]
    przystanki_obsługiwane = {przystanek: False for przystanek in przystanki1}

    for wariant in warianty_kursow:
        # Przejdź przez wszystkie przystanki, które są obsługiwane w danym wariancie
        for przystanek in wariant["zatrzymywane_przystanki"]:
            # Jeśli przystanek jest obsługiwany przez dany wariant, oznacz go jako obsługiwany
            if przystanek in przystanki1:
                przystanki_obsługiwane[przystanek] = True
    
    # Zwracamy listę przystanków, które są obsługiwane przez przynajmniej jeden kurs
    przystanki_po_usunieciu = [przystanek for przystanek, obsługiwany in przystanki_obsługiwane.items() if obsługiwany]

    return przystanki_po_usunieciu



def oblicz_ilosc_wariantow_dla_przystanku(sheet, przystanki,  numer_wiersza):
    warianty_kursow_dla_przystanku = sprawdz_wariant_od_wiersza(sheet, przystanki, numer_wiersza)
    unikalne_warianty = set()  # Tworzymy zbiór, aby przechowywać unikalne warianty
    for wariant_info in warianty_kursow_dla_przystanku:
        unikalne_warianty.add(wariant_info['wariant'])  # Dodajemy wariant, jeśli przystanek nie jest pomijany
    
    ilosc_wariantow = len(unikalne_warianty)  # Liczymy unikalne warianty dla danego przystanku

    return ilosc_wariantow


def oblicz_maks_ilosc_wariantow(sheet, przystanki):
    max_ilosc=0
    for wiersz_startowy in range(2, sheet.max_row + 1):

        warianty_kursow_dla_przystanku = sprawdz_wariant_od_wiersza(sheet, przystanki, wiersz_startowy)
        unikalne_warianty = set()
        for wariant_info in warianty_kursow_dla_przystanku:
            unikalne_warianty.add(wariant_info['wariant'])
        if max_ilosc<len(unikalne_warianty):
            max_ilosc=len(unikalne_warianty)

    return max_ilosc
def przygotuj_warianty(warianty_kursow):
    # Tworzymy listę unikalnych nazw wariantów oraz listę dla wynikowych wariantów
    unikalne_nazwy_wariantow = []
    unikalne_warianty = []

    for wariant_info in warianty_kursow:
        # Sprawdzamy, czy wariant jest już w liście unikalnych nazw
        if wariant_info['wariant'] not in unikalne_nazwy_wariantow:
            # Jeśli wariant nie jest obecny, dodajemy jego nazwę do listy unikalnych nazw
            unikalne_nazwy_wariantow.append(wariant_info['wariant'])
            # Dodajemy pełne informacje o wariancie do unikalnych wariantów
            unikalne_warianty.append({
                "wariant": wariant_info['wariant'],
                "zatrzymywane_przystanki": wariant_info['zatrzymywane_przystanki'],
                "czas_przejazdu_miedzy_przystankami": wariant_info['czas_przejazdu_miedzy_przystankami']
            })

    # Zwracamy listę unikalnych wariantów
    return unikalne_warianty


file_name="Rozkład"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = file_name

#Funkcja automatycznie dopasowanie szerokości kolumny 
def auto_fit_column(worksheet, column_start):
    max_length = 0

    for row in range(1, worksheet.max_row + 1):  # Przechodzimy przez wszystkie wiersze
        cell_value = worksheet[f'{column_start}{row}'].value
        if cell_value:
            max_length = max(max_length, len(str(cell_value)))  # Znajdujemy najdłuższy tekst

    adjusted_width = max_length   # Dodajemy margines
    worksheet.column_dimensions[column_start].width = adjusted_width

def polacz_godziny_odjazdow(kolumny, numer_wiersza,przystanki):
    a=''
    warianty_kursow=sprawdz_wariant_od_wiersza(sheet,przystanki, numer_wiersza)
    for wariant in warianty_kursow:
        if wariant['kolumna'] in kolumny:
            godzina_odjazdu = sheet.cell(numer_wiersza, wariant['kolumna']).value  # Odczytujemy wartość z arkusza
            war=wariant['wariant']
            if war > 'A':  # Sprawdzamy, czy to litera i czy większa niż 'A'
                wari = chr(ord(war) - 1)  # Zmniejszamy literę o jeden w alfabecie
            else:
                wari = ''  # Jeśli wariant to 'A', zmieniamy na pusty string
            if godzina_odjazdu:  # Jeśli godzina jest niepusta
                a += str(godzina_odjazdu)+ wari + '   '  # Dołączamy godzinę do stringa 'a' z odstępem
    return a


def wstaw_rozklad_dla_przystanku(sheet,przystanki, start_row, ws, numer_wiersza):
    max=oblicz_maks_ilosc_wariantow(sheet,przystanki)
    
    ostatnia_kolumna = chr(ord('C') + max)
    ws.merge_cells(f'A{start_row}:{ostatnia_kolumna}{start_row}')
    ws[f'A{start_row}'] = "Rozkład ważny od: 28:06.2021 r."
    ws[f'A{start_row}'].font = header_white_font
    ws[f'A{start_row}'].alignment = alignment_right
    ws[f'A{start_row}'].fill = header_black_fill
    ws[f'A{start_row}'].border=Border(
        left=thick_border.left,
        top=thick_border.top
    )
    ws.merge_cells(f'A{1+start_row}:B{2+start_row}')
    ws[f'A{1+start_row}'] = "Kierunek"
    ws[f'A{1+start_row}'].font = text_font
    ws[f'A{1+start_row}'].alignment = alignment_left_corner
    ws[f'A{1+start_row}'].border=Border(
        left=thick_border.left,
        top=thin_border.top
    )
    ws[f'A{2+start_row}'].border=Border(
        left=thick_border.left,
        bottom=thin_border.bottom
    )

    warianty_kursow = sprawdz_wariant_od_wiersza(sheet, przystanki, numer_wiersza )
    przystanki_po_usunieciu=usun_przystanki_bez_kursow(przystanki,warianty_kursow,numer_wiersza)

    ilosc_wariantow=oblicz_ilosc_wariantow_dla_przystanku(sheet,przystanki,numer_wiersza)
    ws.merge_cells(f'C{1+start_row}:{ostatnia_kolumna}{2+start_row}')
    ws[f'C{1+start_row}'] = przystanki_po_usunieciu[-1]
    
    ws[f'C{1+start_row}'].font = header_black_font
    ws[f'C{1+start_row}'].alignment = alignment_center

    ws.merge_cells(f'A{3+start_row}:A{4+start_row}')
    ws[f'A{3+start_row}'] = 'L.p.'
    ws[f'A{3+start_row}'].fill = header_grey_fill
    ws[f'A{3+start_row}'].alignment = alignment_center
    ws[f'A{3+start_row}'].font = header_black_font

    ws[f'A{3+start_row}'].border=Border(
        left=thick_border.left,
        right=thin_border.right,
        top=thin_border.top
    )
    ws[f'A{4+start_row}'].border=Border(
        left=thick_border.left,
        right=thin_border.right,
        bottom=thin_border.bottom
    )
    kolumna_trasa_i_czas = chr(ord('B') + max-ilosc_wariantow)
    ws.merge_cells(f'B{3+start_row}:{kolumna_trasa_i_czas}{4+start_row}')
    ws[f'B{3+start_row}'] = 'Trasa i czas przejazdu'
    ws[f'B{3+start_row}'].fill = header_grey_fill
    ws[f'B{3+start_row}'].alignment = alignment_center
    ws[f'B{3+start_row}'].font = header_black_font



    is_first_col_empty = True
    alphabet_start = ord('A')  
    start_col=ord('C')
    start_col_num = ord('C') + max - ilosc_wariantow  
    end_col_num = ord('C') + max  
    for col_num in range(start_col, end_col_num):
        col_letter = chr(col_num)
        ws.column_dimensions[f'{col_letter}'].width = 2.75
        ws[f'{col_letter}{5+start_row}'].fill = header_black_fill
        ws[f'{col_letter}{5+start_row}'].font = white_text_font
        ws[f'{col_letter}{5+start_row}'] = 0
    for col_num in range(start_col_num, end_col_num ):
        
        col_letter = chr(col_num)
        
        ws.merge_cells(f'{col_letter}{3+start_row}:{col_letter}{4+start_row}')
        ws[f'{col_letter}{3+start_row}'].fill = header_grey_fill
        ws[f'{col_letter}{3+start_row}'].alignment = alignment_center
        ws[f'{col_letter}{3+start_row}'].font = header_black_font
        ws[f'{col_letter}{3+start_row}'] = f'Kolumna {col_letter}'  
        if is_first_col_empty:
            ws[f'{col_letter}{3+start_row}'] = ""
            is_first_col_empty = False
        else:
            ws[f'{col_letter}{3+start_row}'] = chr(alphabet_start)
            alphabet_start += 1
        
    

    przystanki_po_usunieciu=usun_przystanki_bez_kursow(przystanki, warianty_kursow,numer_wiersza)
    start_row+=5
    unikalne_warianty=przygotuj_warianty(warianty_kursow)

    for index, przystanek in enumerate(przystanki_po_usunieciu, start=1):
        ws[f'A{start_row+index-1}'].border=Border(
        left=thick_border.left
        )
        if index==1:
            ws[f'A{start_row + index - 1}'].fill=header_black_fill
            ws[f'B{start_row + index - 1}'].fill=header_black_fill
            ws[f'A{start_row + index - 1}'].font=header_white_font
            ws[f'B{start_row + index - 1}'].font=header_white_font

        #Numer wiersza

        ws[f'A{start_row + index - 1}'] = f'{index-1}'
        ws[f'A{start_row+index-1}'].border=Border(
        left=thick_border.left
        )
        #Przystanek
        ws[f'B{start_row + index - 1}'] = przystanek
        ws.merge_cells(f'B{start_row+index-1}:{kolumna_trasa_i_czas}{start_row+index-1}')

    # Wstawianie czasu przejazdu dla każdego wariantu
        i=0
        for wariant_info in unikalne_warianty:
            #print(wariant_info["wariant"])
            for czas in wariant_info["czas_przejazdu_miedzy_przystankami"]:
                ws[f'{chr(start_col_num + i)}{start_row + index - 1}'].border=thin_border
        # Sprawdzamy, czy przystanek końcowy zgadza się z bieżącym przystankiem
                if czas["do_przystanku"] == przystanek:
                    czas_przejazdu = czas["czas_przejazdu"]
                    
                    if czas_przejazdu is not None:
                # Wstawiamy zaokrąglony czas przejazdu w odpowiedniej kolumnie
                        ws[f'{chr(start_col_num + i)}{start_row + index - 1}'] = round(czas_przejazdu)
            i+=1
    #ws[f'B{4+start_row}'].border=thin_border
    #ws[f'C{4+start_row}'].border=thin_border
    #ws[f'B{5+start_row}'].border=thin_border
    #ws[f'C{5+start_row}'].border=thin_border
    
    #Tutaj zaczyna się część związana z godzinami odjazdów


    index_linie_godziny = -2
    if dni_robocze_szkolne:
        ws.merge_cells(f'{ostatnia_kolumna}{index_linie_godziny+start_row}:{ostatnia_kolumna}{index_linie_godziny+1+start_row}')
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'] = 'Dni robocze w dni nauki szkolnej'
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].fill = header_dniszkolne_fill
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].alignment = alignment_center
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].font = header_black_font

        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].border=thin_border
        ws[f'{ostatnia_kolumna}{index_linie_godziny+1+start_row}'].border=Border(
            left=thick_border.left,
            right=thick_border.right,
            bottom=thin_border.bottom,
            top=thin_border.top)

        ws.merge_cells(f'{ostatnia_kolumna}{index_linie_godziny+2+start_row}:{ostatnia_kolumna}{index_linie_godziny+3+start_row}')
        
        ws[f'{ostatnia_kolumna}{index_linie_godziny+2+start_row}'] = polacz_godziny_odjazdow(dni_robocze_szkolne,numer_wiersza,przystanki)
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row+2}'].alignment = Alignment(wrap_text=True)
        ws[f'{ostatnia_kolumna}{index_linie_godziny+2+start_row}'].alignment = alignment_hour
        ws[f'{ostatnia_kolumna}{index_linie_godziny+2+start_row}'].font = text_font

        ws[f'{ostatnia_kolumna}{index_linie_godziny+2+start_row}'].border=thin_border
        ws[f'{ostatnia_kolumna}{index_linie_godziny+3+start_row}'].border=thin_border
        index_linie_godziny+=4
    if wakacje:
        ws.merge_cells(f'{ostatnia_kolumna}{index_linie_godziny+start_row}:{ostatnia_kolumna}{index_linie_godziny+start_row+1}')
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'] = 'Dni robocze wolne od nauki szkolnej'
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].fill = header_wakacje_fill
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].alignment = alignment_center
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].font = header_black_font

        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].border=Border(
            left=thick_border.left,
            right=thick_border.right,
            bottom=thin_border.bottom,
            top=thin_border.top)
        ws[f'{ostatnia_kolumna}{index_linie_godziny+1+start_row}'].border=Border(
            left=thick_border.left,
            right=thick_border.right,
            bottom=thin_border.bottom,
            top=thin_border.top)
        index_linie_godziny+=2
        ws.merge_cells(f'{ostatnia_kolumna}{index_linie_godziny+start_row}:{ostatnia_kolumna}{index_linie_godziny+1+start_row}')
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'] = polacz_godziny_odjazdow(wakacje,numer_wiersza,przystanki)
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].alignment = Alignment(wrap_text=True)
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].alignment = alignment_hour
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].font = text_font

        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].border=Border(
            left=thick_border.left,
            right=thick_border.right,
            bottom=thin_border.bottom,
            top=thin_border.top)
        ws[f'{ostatnia_kolumna}{index_linie_godziny+1+start_row}'].border=Border(
            left=thick_border.left,
            right=thick_border.right,
            bottom=thin_border.bottom,
            top=thin_border.top)
        index_linie_godziny+=2
    if soboty:
        ws.merge_cells(f'{ostatnia_kolumna}{index_linie_godziny+start_row}:{ostatnia_kolumna}{index_linie_godziny+1+start_row}')
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'] = 'Soboty'
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].fill = header_wakacje_fill
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].alignment = alignment_center
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].font = header_black_font

        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].border=Border(
            left=thick_border.left,
            right=thick_border.right,
            bottom=thin_border.bottom,
            top=thin_border.top)
        ws[f'{ostatnia_kolumna}{index_linie_godziny+1+start_row}'].border=Border(
            left=thick_border.left,
            right=thick_border.right,
            bottom=thin_border.bottom,
            top=thin_border.top)

        index_linie_godziny+=2
        ws.merge_cells(f'{ostatnia_kolumna}{index_linie_godziny+start_row}:{ostatnia_kolumna}{index_linie_godziny+1+start_row}')
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'] = polacz_godziny_odjazdow(soboty,numer_wiersza,przystanki)
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].alignment = Alignment(wrap_text=True)
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].alignment = alignment_hour
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].font = text_font

        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].border=Border(
            left=thick_border.left,
            right=thick_border.right,
            bottom=thin_border.bottom,
            top=thin_border.top)
        ws[f'{ostatnia_kolumna}{index_linie_godziny+1+start_row}'].border=Border(
            left=thick_border.left,
            right=thick_border.right,
            bottom=thin_border.bottom,
            top=thin_border.top)
        index_linie_godziny+=2
    if niedziele:
        ws.merge_cells(f'{ostatnia_kolumna}{index_linie_godziny+start_row}:{ostatnia_kolumna}{index_linie_godziny+1+start_row}')
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'] = 'Niedziele'
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].fill = header_wakacje_fill
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].alignment = alignment_center
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].font = header_black_font

        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].border=Border(
            left=thick_border.left,
            right=thick_border.right,
            bottom=thin_border.bottom,
            top=thin_border.top)
        ws[f'{ostatnia_kolumna}{index_linie_godziny+1+start_row}'].border=Border(
            left=thick_border.left,
            right=thick_border.right,
            bottom=thin_border.bottom,
            top=thin_border.top)
        index_linie_godziny+=2

        ws.merge_cells(f'{ostatnia_kolumna}{index_linie_godziny+start_row}:{ostatnia_kolumna}{index_linie_godziny+1+start_row}')
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'] = polacz_godziny_odjazdow(niedziele,numer_wiersza,przystanki)
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].alignment = Alignment(wrap_text=True)
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].alignment = alignment_hour
        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].font = text_font

        ws[f'{ostatnia_kolumna}{index_linie_godziny+start_row}'].border=Border(
            left=thick_border.left,
            right=thick_border.right,
            bottom=thin_border.bottom,
            top=thin_border.top)
        ws[f'{ostatnia_kolumna}{index_linie_godziny+1+start_row}'].border=Border(
            left=thick_border.left,
            right=thick_border.right,
            bottom=thin_border.bottom,
            top=thin_border.top)
        index_linie_godziny+=2
    start_row+= 6



wstaw_rozklad_dla_przystanku(sheet,przystanki, 1, ws,5)

ws.column_dimensions['A'].width =4.14
auto_fit_column(ws,'B')


wb.save("rozkład.xlsx")

print("Plik rozkład_jazdy.xlsx został zapisany.")
